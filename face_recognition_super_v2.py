
import os, sys, time, threading, queue, json, pickle
from datetime import datetime, timedelta, time as dtime
from typing import List, Tuple, Dict, Any, Optional
from pathlib import Path

import cv2
import numpy as np
import torch
import openpyxl

from ultralytics import YOLO
from facenet_pytorch import MTCNN, InceptionResnetV1

try:
    import faiss
    FAISS_AVAILABLE = True
except Exception:
    FAISS_AVAILABLE = False

YOLO_MODEL_PATH = 'detection/weights/yolov11n-face.pt'
EMBEDDINGS_FILE = 'known_embeddings.pkl'
EXCEL_FILE = 'DashBoard.xlsx'
STREAM_URL = 'rtsp://admin:1234qwer@@172.16.0.3:554/Streaming/Channels/102?tcp'
OUTPUT_DIR = 'recordings'
OUTPUT_FPS = 30.0

RECOGNITION_THRESHOLD = 0.9
YOLO_CONF = 0.45

MAX_BOX_DISTANCE = 60
CONFIRM_FRAMES = 2
REQUIRED_VOTES = 1
EMBEDDING_TTL_SECONDS = 1.0

FRAME_Q_MAX = 1
VIDEO_Q_MAX = 512
EXCEL_Q_MAX = 1024

os.makedirs(OUTPUT_DIR, exist_ok=True)

DEVICE = "cpu"
print(f"[System] Using device: {DEVICE} (cuda available: {torch.cuda.is_available()})")

def now_str():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

def parse_time(s: str) -> dtime:
    try:
        if s is None: return dtime(hour=0, minute=0)
        s_val = str(s).strip()
        if ':' in s_val:
            parts = s_val.split(':'); h = int(parts[0]) if parts[0] != '' else 0
            m = int(parts[1]) if len(parts) > 1 and parts[1] != '' else 0
        else:
            h = int(s_val); m = 0
        h = max(0, min(23, h)); m = max(0, min(59, m))
        return dtime(hour=h, minute=m)
    except Exception:
        return dtime(hour=0, minute=0)

def load_times(config_path='config.json'):
    defaults = {'start_time': '06:00', 'lateness_time': '07:00', 'absence_time': '10:00'}
    try:
        if os.path.exists(config_path):
            with open(config_path, 'r') as f:
                cfg = json.load(f)
            merged = defaults.copy(); merged.update({k: v for k,v in cfg.items() if k in merged})
        else:
            merged = defaults
    except Exception:
        merged = defaults
    return parse_time(merged.get('start_time')), parse_time(merged.get('lateness_time')), parse_time(merged.get('absence_time'))

def is_time_in_range(start: dtime, end: dtime, now_t: dtime) -> bool:
    if start <= end:
        return start <= now_t < end
    else:
        return now_t >= start or now_t < end

print("[System] Loading models...")
try:
    model_yolo = YOLO(YOLO_MODEL_PATH)
    if DEVICE == 'cuda':
        try:
            model_yolo.to('cuda')
        except Exception:
            pass

    print("[System] YOLO loaded.")
except Exception as e:
    print(f"[ERROR] YOLO load failed: {e}")
    model_yolo = None

try:
    mtcnn = MTCNN(keep_all=False, image_size=160, device=DEVICE)
    resnet = InceptionResnetV1(pretrained='vggface2').eval().to(DEVICE)
    print("[System] MTCNN and ResNet loaded.")
except Exception as e:
    print(f"[ERROR] Face models load failed: {e}")
    mtcnn = None; resnet = None

FAISS_INDEX = None
KNOWN_NAMES: List[str] = []
KNOWN_EMBEDDINGS = None

def load_known_embeddings(path=EMBEDDINGS_FILE):
    global FAISS_INDEX, KNOWN_NAMES, KNOWN_EMBEDDINGS
    if not os.path.exists(path):
        print(f"[WARN] Embeddings file not found: {path}")
        return False
    try:
        with open(path, 'rb') as f:
            raw = pickle.load(f)
        flat_names = []; flat_embs = []
        for name, arr in raw.items():
            for e in arr:
                flat_names.append(name); flat_embs.append(np.asarray(e, dtype='float32'))
        if not flat_embs:
            print("[WARN] No embeddings available")
            return False
        mat = np.vstack(flat_embs).astype('float32')
        KNOWN_EMBEDDINGS = mat
        if FAISS_AVAILABLE:
            D = mat.shape[1]; idx = faiss.IndexFlatL2(D); idx.add(mat); FAISS_INDEX = idx
            print(f"[INFO] FAISS index built: vectors={mat.shape[0]}, dim={D}")
        else:
            FAISS_INDEX = None
            print("[INFO] Faiss not available; using NumPy fallback")
        KNOWN_NAMES = flat_names
        return True
    except Exception as e:
        print(f"[ERROR] load_known_embeddings: {e}")
        return False


def compute_embeddings_batch(crops: List[np.ndarray]) -> List[Optional[np.ndarray]]:
    if not crops: return []
    valid_pils = []; idx_map = []; fixed_size=(160,160)
    for i, c in enumerate(crops):
        try:
            if c is None or not isinstance(c, np.ndarray) or c.size==0:
                continue
            resized = cv2.resize(c, fixed_size)
            rgb = cv2.cvtColor(resized, cv2.COLOR_BGR2RGB)
            from PIL import Image
            pil = Image.fromarray(rgb)
            valid_pils.append(pil); idx_map.append(i)
        except Exception:
            continue
    if not valid_pils:
        return [None]*len(crops)
    try:
        with torch.no_grad():
            aligned = mtcnn(valid_pils)
            if aligned is None:
                return [None]*len(crops)
            if isinstance(aligned, list):
                aligned = [t for t in aligned if t is not None]
                if not aligned:
                    return [None]*len(crops)
                aligned = torch.stack(aligned)
            aligned = aligned.to(DEVICE)
            embs_t = resnet(aligned)
            embs = embs_t.detach().cpu().numpy()
    except Exception as e:
        print(f"[ERROR] compute_embeddings_batch: {e}")
        return [None]*len(crops)
    result = [None]*len(crops)
    for out_idx, emb in zip(idx_map, embs):
        e = np.asarray(emb, dtype='float32')
        if e.shape[0] != 512:
            if e.shape[0] > 512:
                e = e[:512]
            else:
                e = np.pad(e, (0, 512-e.shape[0])).astype('float32')
        result[out_idx] = e
    return result

frame_q = queue.Queue(maxsize=FRAME_Q_MAX)
video_q = queue.Queue(maxsize=VIDEO_Q_MAX)
excel_q = queue.Queue(maxsize=EXCEL_Q_MAX)

recognized_names = set()

class CameraThread(threading.Thread):
    def __init__(self, src: str, frame_q: queue.Queue):
        super().__init__(daemon=True)
        self.src = src; self.frame_q = frame_q
        self.cap = cv2.VideoCapture(src)
        self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        self.running = False
        if not self.cap.isOpened():
            print(f"[ERROR] Camera open failed: {src}")
        else:
            print("[INFO] Camera opened successfully")
            self.running = True
    def run(self):
        dropped = 0
        while self.running:
            ret, frame = self.cap.read()
            if not ret or frame is None:
                dropped += 1
                if dropped % 50 == 0:
                    print(f"[Camera] No frame returned for {dropped} attempts")
                time.sleep(0.05); continue
            dropped = 0
            ts = datetime.now()
            try:
                self.frame_q.put_nowait((frame.copy(), ts))
            except queue.Full:
                try:
                    _ = self.frame_q.get_nowait()
                    self.frame_q.put_nowait((frame.copy(), ts))
                except Exception:
                    pass
    def stop(self):
        self.running = False
        try: self.cap.release()
        except: pass
        print("[INFO] CameraThread stopped")



class VideoWriterThread(threading.Thread):
    def __init__(self, video_q: queue.Queue, filename: str, fps: float, frame_size: Tuple[int,int]):
        super().__init__(daemon=True)
        self.video_q = video_q
        self.filename = filename
        self.fps = fps
        self.frame_size = frame_size
        self.running = True

        self._fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        self._writer = cv2.VideoWriter(self.filename, self._fourcc, self.fps, self.frame_size)

        self._frame_count = 0
        self._start_time = time.time()
        self._last_frame_time = None

        print(f"[Video] Recording in real time: {self.filename} ({self.frame_size[0]}x{self.frame_size[1]}, {self.fps} fps)")

    def run(self):
        last_log = time.time()
        while self.running:
            try:
                frame, meta = self.video_q.get(timeout=0.5)
            except queue.Empty:
                continue

            if frame is None:
                continue

            if (frame.shape[1], frame.shape[0]) != self.frame_size:
                frame = cv2.resize(frame, self.frame_size)

            frame_time = meta.get('time', datetime.now())
            if self._last_frame_time is not None:
                diff = (frame_time - self._last_frame_time).total_seconds()
                if diff > 0:
                    time.sleep(diff)
            self._last_frame_time = frame_time

            try:
                self._writer.write(frame)
                self._frame_count += 1
                self.video_q.task_done()
            except Exception as e:
                print(f"[ERROR] Writing frame: {e}")

            if time.time() - last_log > 5.0:
                elapsed = time.time() - self._start_time
                avg_fps = self._frame_count / elapsed if elapsed > 0 else 0
                print(f"[Video] Frames written: {self._frame_count}, Avg FPS: {avg_fps:.2f}")
                last_log = time.time()

        try:
            self._writer.release()
            print(f"[Video] Recording saved: {self.filename}")
        except Exception as e:
            print(f"[ERROR] Closing writer: {e}")

        print("[INFO] VideoWriterThread stopped.")

    def stop(self):
        self.running = False


class ExcelLoggerThread(threading.Thread):
    def __init__(self, excel_q: queue.Queue, path=EXCEL_FILE):
        super().__init__(daemon=True)
        self.excel_q = excel_q; self.path = path; self.running = True
        self._wb = None; self._ws = None; self._name_row = {}
        self._last_save = time.time(); self._save_interval = 30; self.absence_done = False
        self._shutdown_callback = None
        self._init_workbook()
    def _init_workbook(self):
        try:
            if os.path.exists(self.path):
                self._wb = openpyxl.load_workbook(self.path); self._ws = self._wb.active
                print(f"[Excel] Loaded workbook: {self.path}")
            else:
                self._wb = openpyxl.Workbook(); self._ws = self._wb.active; self._ws.title='Attendance'; self._ws.cell(row=1,column=1).value='Students Names'; self._wb.save(self.path); print(f"[Excel] Created workbook: {self.path}")
            self._build_map()
        except Exception as e:
            print(f"[ERROR] Excel init: {e}"); self._wb = None; self._ws = None
    def _build_map(self):
        self._name_row.clear()
        for r in range(2, self._ws.max_row+1):
            nm = self._ws.cell(r,1).value
            if nm: self._name_row[str(nm).upper().strip()] = r
        print(f"[Excel] Cached {len(self._name_row)} names from sheet")
    def set_shutdown_callback(self, cb):
        self._shutdown_callback = cb
    def write_record(self, name, status):
        if self._ws is None: return
        if name in (None,'Unknown','Unknown_Unstable'): return
        nup = str(name).upper().strip()
        if nup not in self._name_row:
            print(f"[Excel] Name not found in sheet: {name} (skipping)"); return
        r = self._name_row[nup]; c = 2
        while self._ws.cell(r,c).value is not None and c<200:
            c += 1
        self._ws.cell(r,c).value = f"{now_str()} - {status}"
        try:
            self._wb.save(self.path); print(f"[Excel] Logged: {name} - {status}")
        except Exception as e:
            print(f"[ERROR] Excel save: {e}")
    def process_absence(self):
        if not KNOWN_NAMES:
            print("[Excel] No KNOWN_NAMES to process absence"); return
        absent = 0; not_found=0; errors=0
        for name in set(KNOWN_NAMES):
            try:
                if name not in recognized_names:
                    nup = str(name).upper().strip()
                    if nup in self._name_row:
                        r = self._name_row[nup]; c=2
                        while self._ws.cell(r,c).value is not None and c<200: c+=1
                        self._ws.cell(r,c).value = f"{now_str()} - Absent"; 
                    else:
                        not_found += 1
            except Exception as e:
                errors+=1; print(f"[ERROR] absence for {name}: {e}")
        try:
            self._wb.save(self.path); print(f"[Excel] Absence saved. Absent={absent}, not_found={not_found}, errors={errors}")
        except Exception as e:
            print(f"[ERROR] Excel final save: {e}")
        self.absence_done = True
        if self._shutdown_callback: self._shutdown_callback()
    def run(self):
        while self.running:
            try:
                task = self.excel_q.get(timeout=0.5)
            except queue.Empty:
                if time.time() - self._last_save >= self._save_interval and self._wb:
                    try: self._wb.save(self.path); self._last_save = time.time(); print("[Excel] Auto-saved") 
                    except Exception as e: print(f"[ERROR] autosave: {e}")
                continue
            if task is None:
                break
            ttype, payload = task
            if ttype == 'LOG':
                name, status = payload; self.write_record(name, status)
            elif ttype == 'ABSENCE':
                print('[Excel] Received ABSENCE task'); self.process_absence()
            try: self.excel_q.task_done()
            except: pass
        print('[Excel] Exiting Excel thread')
    def stop(self):
        self.running = False
        try:
            if self._wb: self._wb.save(self.path)
        except Exception as e:
            print(f"[ERROR] Excel stop save: {e}")


class RecognitionThread(threading.Thread):
    def __init__(self, frame_q: queue.Queue, video_q: queue.Queue, excel_q: queue.Queue):
        super().__init__(daemon=True)
        self.frame_q = frame_q
        self.video_q = video_q
        self.excel_q = excel_q
        self.running = True
        self.tracks = {}
        self.next_id = 0
        from concurrent.futures import ThreadPoolExecutor
        self.pool = ThreadPoolExecutor(max_workers=2)
        self.start_time_obj, self.lateness_time_obj, self.absence_time_obj = load_times()
        self._fps_prev = time.time()
        self._fps_count = 0
        self.realtime_fps = 0.0

        self.faiss_index = None
        self.track_ids_for_index = []
        self.embedding_dim = None

    def update_faiss_index(self):
        emb_list = []
        self.track_ids_for_index = []
        for tid, t in self.tracks.items():
            emb = t.get('cached_embedding')
            if emb is not None:
                emb_list.append(emb.astype('float32'))
                self.track_ids_for_index.append(tid)
        if not emb_list:
            self.faiss_index = None
            self.embedding_dim = None
            return
        emb_matrix = np.vstack(emb_list)
        self.embedding_dim = emb_matrix.shape[1]
        index = faiss.IndexFlatL2(self.embedding_dim)
        index.add(emb_matrix)
        self.faiss_index = index

    def find_track_by_embedding(self, emb, threshold=0.6):
        if self.faiss_index is None or emb.shape[0] != self.embedding_dim:
            return None
        query = emb.reshape(1, -1).astype('float32')
        D, I = self.faiss_index.search(query, 1)
        dist = float(D[0][0])
        idx = int(I[0][0])
        if dist < (threshold ** 2):
            return self.track_ids_for_index[idx]
        return None

    def stable_name(self, hist):
        if not hist:
            return 'Unknown_Unstable'
        counts = {}
        for n in hist:
            counts[n] = counts.get(n, 0) + 1
        best = max(counts, key=counts.get)
        if counts[best] >= REQUIRED_VOTES:
            return best
        return 'Unknown_Unstable'

    def _update_fps(self):
        self._fps_count += 1
        now = time.time()
        elapsed = now - self._fps_prev
        if elapsed >= 1.0:
            self.realtime_fps = self._fps_count / elapsed
            self._fps_count = 0
            self._fps_prev = now

    def run(self):
        global KNOWN_NAMES, FAISS_INDEX, recognized_names
        print('[Recognition] Started')

        while self.running:
            try:
                frame, ts = self.frame_q.get(timeout=0.5)
            except queue.Empty:
                continue

            self._update_fps()
            H, W = frame.shape[:2]
            nowt = ts.time()

            if not is_time_in_range(self.start_time_obj, self.absence_time_obj, nowt):
                try:
                    self.video_q.put_nowait((frame.copy(), {'time': datetime.now()}))
                except queue.Full:
                    pass
                try:
                    self.frame_q.task_done()
                except:
                    pass
                continue

            try:
                yres = model_yolo(frame, verbose=False, conf=YOLO_CONF)
            except Exception as e:
                print(f'[WARN] YOLO call failed: {e}')
                yres = None

            detections = []
            if yres is not None and len(yres) > 0:
                for b in yres[0].boxes:
                    try:
                        x1, y1, x2, y2 = map(int, b.xyxy[0].tolist())
                    except Exception:
                        vals = b.xyxy[0].cpu().numpy()
                        x1, y1, x2, y2 = map(int, vals.tolist())
                    detections.append((x1, y1, x2, y2))

            analyze_crops = []
            analyze_tids = []
            new_tracks = {}

            self.update_faiss_index()

            for box in detections:
                x1, y1, x2, y2 = box
                x1, y1, x2, y2 = max(0, x1), max(0, y1), min(W - 1, x2), min(H - 1, y2)
                crop = frame[y1:y2, x1:x2]

                temp_emb = None
                if crop is not None and crop.size > 0:
                    try:
                        temp_emb_list = compute_embeddings_batch([crop])
                        temp_emb = temp_emb_list[0] if temp_emb_list else None
                    except Exception as e:
                        print(f'[WARN] Temp embedding failed: {e}')
                        temp_emb = None

                tid = None
                if temp_emb is not None:
                    tid = self.find_track_by_embedding(temp_emb)

                if tid is None:
                    tid = self.next_id
                    self.next_id += 1
                    self.tracks[tid] = {
                        'box': (x1, y1, x2, y2),
                        'history': [],
                        'seen_count': 0,
                        'analysis_runs': 0,
                        'last_recognition': 'Unknown',
                        'cached_embedding': None,
                        'cached_time': None,
                        'logged': False,
                        'first_seen_time': None,
                        'last_seen_time': None
                    }

                t = self.tracks[tid]
                t['box'] = (x1, y1, x2, y2)
                t['seen_count'] = t.get('seen_count', 0) + 1

                use_cached = False
                cached_time = t.get('cached_time')
                if t.get('cached_embedding') is not None and cached_time:
                    if (datetime.now() - cached_time).total_seconds() <= EMBEDDING_TTL_SECONDS:
                        use_cached = True

                analyze_now = False
                sc = t['seen_count']
                ar = t.get('analysis_runs', 0)
                if sc < 3:
                    analyze_now = True
                else:
                    if ar in (0, 1):
                        analyze_now = True

                if use_cached:
                    rec = t.get('last_recognition', 'Unknown')
                    t['history'] = (t['history'] + [rec])[-CONFIRM_FRAMES:]
                elif analyze_now:
                    if crop is None or crop.size == 0:
                        t['history'] = (t['history'] + ['Unknown'])[-CONFIRM_FRAMES:]
                    else:
                        analyze_crops.append(crop)
                        analyze_tids.append(tid)
                else:
                    rec = t.get('last_recognition', 'Unknown')
                    t['history'] = (t['history'] + [rec])[-CONFIRM_FRAMES:]

                t['last_time_seen'] = datetime.now()
                new_tracks[tid] = t

            if analyze_crops:
                future = self.pool.submit(compute_embeddings_batch, analyze_crops)
                try:
                    embeddings = future.result(timeout=3.0)
                except Exception as e:
                    print(f'[WARN] embedding batch failed/timeout: {e}')
                    embeddings = [None] * len(analyze_crops)

                valid_embs = []
                valid_tids = []
                for i, emb in enumerate(embeddings):
                    tid = analyze_tids[i]
                    if emb is None:
                        new_tracks[tid]['history'] = (new_tracks[tid]['history'] + ['Unknown'])[-CONFIRM_FRAMES:]
                        new_tracks[tid]['analysis_runs'] = new_tracks[tid].get('analysis_runs', 0) + 1
                        new_tracks[tid]['last_recognition'] = 'Unknown'
                        continue
                    new_tracks[tid]['cached_embedding'] = emb
                    new_tracks[tid]['cached_time'] = datetime.now()
                    new_tracks[tid]['analysis_runs'] = new_tracks[tid].get('analysis_runs', 0) + 1
                    valid_embs.append(emb)
                    valid_tids.append(tid)

                if valid_embs:
                    qmat = np.vstack(valid_embs).astype('float32')
                    if FAISS_INDEX is not None:
                        D, I = FAISS_INDEX.search(qmat, 1)
                    else:
                        A = KNOWN_EMBEDDINGS
                        D_list = []; I_list = []
                        for q in qmat:
                            dists = np.linalg.norm(A - q, axis=1)
                            idx = np.argmin(dists)
                            D_list.append([dists[idx]]); I_list.append([idx])
                        D = np.array(D_list); I = np.array(I_list)

                    for vi, tid in enumerate(valid_tids):
                        dist = float(D[vi][0])
                        idx = int(I[vi][0])
                        if dist < (RECOGNITION_THRESHOLD ** 2) and idx < len(KNOWN_NAMES):
                            name = KNOWN_NAMES[idx]
                        else:
                            name = 'Unknown'
                        new_tracks[tid]['last_recognition'] = name
                        new_tracks[tid]['history'] = (new_tracks[tid]['history'] + [name])[-CONFIRM_FRAMES:]

            for tid, t in new_tracks.items():
                x1, y1, x2, y2 = t['box']
                hist = t.get('history', [])
                stable = self.stable_name(hist)
                display = t.get('last_recognition', 'Unknown')
                color = (0, 0, 255)
                status = 'Unknown'

                nowt = datetime.now().time()
                if is_time_in_range(self.start_time_obj, self.lateness_time_obj, nowt):
                    status = 'Present'
                    color = (0, 255, 0)
                elif is_time_in_range(self.lateness_time_obj, self.absence_time_obj, nowt):
                    status = 'Late'
                    color = (0, 140, 255)

                if stable == 'Unknown_Unstable' or display == 'Unknown':
                    cur_time = datetime.now().strftime('%H:%M:%S')
                    label_line1 = f"Unknown - {status} | {cur_time}"
                    cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    y_text = max(12, y1 - 25)
                    cv2.putText(frame, label_line1, (x1, y_text),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 255), 2)
                    continue

                display = stable

                if not t.get('logged', False):
                    t['first_seen_time'] = datetime.now()
                    t['first_status'] = status
                    t['first_color'] = color
                    t['logged'] = True
                    try:
                        if display not in recognized_names:
                            self.excel_q.put_nowait(('LOG', (display, status)))
                            recognized_names.add(display)
                    except queue.Full:
                        pass

                t['last_seen_time'] = datetime.now()
                t['last_status'] = status

                first_time_str = t['first_seen_time'].strftime('%H:%M:%S') if t.get('first_seen_time') else ''
                last_time_str = t['last_seen_time'].strftime('%H:%M:%S') if t.get('last_seen_time') else ''

                label_line1 = f"{display} - {status} | {last_time_str}"
                label_line2 = f"{t.get('first_status', 'Unknown')} | {first_time_str}"

                first_color = color
                sec_status = t.get('first_status', 'Unknown')
                if sec_status == 'Present':
                    second_color = (0, 255, 0)
                elif sec_status == 'Late':
                    second_color = (0, 140, 255)
                else:
                    second_color = (255, 255, 255)

                cv2.rectangle(frame, (x1, y1), (x2, y2), first_color, 2)
                y_text = max(12, y1 - 25)
                cv2.putText(frame, label_line1, (x1, y_text),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.7, first_color, 2)
                cv2.putText(frame, label_line2, (x1, y_text + 22),
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, second_color, 2)

            fps_text = f"FPS: {self.realtime_fps:.2f}"
            cv2.putText(frame, fps_text, (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 0), 2)

            for tid, new_t in new_tracks.items():
                if tid in self.tracks:
                    old_t = self.tracks[tid]
                    if old_t.get('logged'):
                        new_t['first_seen_time'] = old_t.get('first_seen_time')
                        new_t['first_status'] = old_t.get('first_status')
                        new_t['first_color'] = old_t.get('first_color')
                        new_t['logged'] = True
                    self.tracks[tid].update(new_t)
                else:
                    self.tracks[tid] = new_t

            try:
                self.video_q.put_nowait((frame.copy(), {'time': datetime.now()}))
            except queue.Full:
                pass
            try:
                self.frame_q.task_done()
            except:
                pass

            cur_time = datetime.now().time()
            if cur_time >= self.absence_time_obj:
                print('[Recognition] Absence time reached, sending ABSENCE task to Excel thread')
                try:
                    self.excel_q.put_nowait(('ABSENCE', None))
                except queue.Full:
                    print('[WARN] Excel queue full when sending ABSENCE task')

        print('[Recognition] Stopped')

    def stop(self):
        self.running = False
        try:
            self.pool.shutdown(wait=False)
        except:
            pass


threads = {}
cam = None; vw = None; xl = None; rec = None

def shutdown_system():
    print('[System] Shutdown initiated...')
    try:
        if rec and getattr(rec,'is_alive',False): rec.stop()
    except: pass
    try:
        if vw and getattr(vw,'running',False):
            vw.stop()
    except: pass
    try:
        if xl and getattr(xl,'running',False): xl.stop()
    except: pass
    try:
        if cam and getattr(cam,'running',False): cam.stop()
    except: pass
    time.sleep(1)
    try:
        cv2.destroyAllWindows(); cv2.waitKey(1)
    except: pass
    print('[System] Shutdown complete. Exiting.')
    sys.exit(0)
def main():
    global cam, vw, xl, rec, KNOWN_NAMES, FAISS_INDEX
    _,_,absence_time_obj = load_times()
    cur_t = datetime.now().time()
    if cur_t >= absence_time_obj :
        print('[Main]Absence Time Reached...')
        try:
            shutdown_system()
            if rec and getattr(rec,'running',False): rec.stop()
            if vw and getattr(vw,'running',False): vw.stop()
            if xl and getattr(xl,'running',False): xl.stop()
            if cam and getattr(cam,'running',False): cam.stop()
            time.sleep(1)
            cv2.destroyAllWindows(); cv2.waitKey(1)
        except Exception as e:
            print(f'[ERROR] final cleanup: {e}')
        print('[System] Exited.')


    print('[System] Starting SmartScan Super v2 Debug')
    ok = load_known_embeddings()
    if not ok:
        print('[WARN] No embeddings loaded; recognition will be Unknown')
    else:
        print(f'[System] Loaded {len(KNOWN_NAMES)} known identities')
    frame_q_local = frame_q
    cam = CameraThread(STREAM_URL, frame_q_local); cam.start()
    first_frame = None; t0=time.time()
    while True:
        try:
            first_frame, ts = frame_q_local.get(timeout=5.0); frame_q_local.task_done(); break
        except queue.Empty:
            if time.time()-t0 > 15:
                print('[ERROR] No frames from camera within 15s. Exiting.'); cam.stop(); return
            print('[System] Waiting for first frame...'); time.sleep(1)
    H,W = first_frame.shape[:2]
    filename = os.path.join(OUTPUT_DIR, f"record_{datetime.now().strftime('%Y%m%d_%H%M%S')}.mp4")
    vw = VideoWriterThread(video_q, filename, OUTPUT_FPS, (W,H));vw.start()
    xl = ExcelLoggerThread(excel_q, EXCEL_FILE); xl.set_shutdown_callback(shutdown_system); xl.start()
    rec = RecognitionThread(frame_q_local, video_q, excel_q); rec.start()
    threads.update({'cam':cam,'vw':vw,'xl':xl,'rec':rec})
    print('[System] All threads started. Viewer loop running. Press q to quit.')
    absence_triggered = False
    _,_,absence_time_obj = load_times()
    try:
        while True:
            try:
                frame, meta = video_q.get(timeout=0.5)
                if frame is not None:
                    cv2.imshow('SmartScan Super v2 Debug', frame)
                try: video_q.task_done()
                except: pass
            except queue.Empty:
                pass
            cur_t = datetime.now().time()
            if cur_t >= absence_time_obj and not absence_triggered:
                print(f'[Main] Absence time reached: {absence_time_obj} - sending ABSENCE task')
                try: excel_q.put(('ABSENCE', None), timeout=5.0); absence_triggered=True
                except queue.Full: print('[ERROR] Could not send ABSENCE task - excel queue full'); shutdown_system(); break
            if cv2.waitKey(1) & 0xFF == ord('q'):
                print('[Main] Manual quit requested'); shutdown_system(); break
            if xl and getattr(xl, 'absence_done', False):
                print('[Main] Excel reported absence done; exiting main loop'); break
    except KeyboardInterrupt:
        print('[Main] KeyboardInterrupt received'); shutdown_system()
    print('[Main] Final cleanup...')
    try:
        if rec and getattr(rec,'running',False): rec.stop()
        if vw and getattr(vw,'running',False): vw.stop()
        if xl and getattr(xl,'running',False): xl.stop()
        if cam and getattr(cam,'running',False): cam.stop()
        time.sleep(1)
        cv2.destroyAllWindows(); cv2.waitKey(1)
    except Exception as e:
        print(f'[ERROR] final cleanup: {e}')
    print('[System] Exited.')

if __name__ == '__main__':
    main()