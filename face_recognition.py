import openpyxl
import os
import cv2
import time
import threading # Required for the CameraHandler class
import torch
import faiss # New: FAISS for fast nearest-neighbor search
# Importing facenet_pytorch components
from facenet_pytorch import MTCNN, InceptionResnetV1
from ultralytics import YOLO
import pickle
import numpy as np
from typing import Dict, Any, List, Tuple, Optional
from datetime import datetime, time as dtime, timedelta
import json

# --- Configuration ---
# NOTE: YOU MUST HAVE THESE WEIGHTS DOWNLOADED AND EMBEDDINGS CREATED
YOLO_MODEL_PATH = 'detection/weights/yolov12n-best.pt' # YOLOv8 model for initial face detection
EMBEDDINGS_FILE = 'known_embeddings.pkl'  # File containing known face embeddings
STREAM_URL = 'rtsp://admin:1234qwer@@172.16.0.3:554/Streaming/Channels/102?tcp' # <<< CHANGE THIS TO YOUR IP CAMERA URL
RECOGNITION_THRESHOLD = 1.0         # Max Euclidean distance for a match (typically 0.9 to 1.1 for FaceNet)
# ADJUSTED: Raised the threshold back up to 0.65. This filters out the false positives 
# (non-face objects) while still being sensitive enough for most faces.
YOLO_CONFIDENCE_THRESHOLD = 0.75       # Minimum confidence score (0.0 to 1.0) for YOLO to accept a detection.

# --- Temporal Consistency Settings (The "Double-Check" Logic) ---
CONFIRM_FRAMES = 1              # Number of recent frames to check for consistency.
REQUIRED_VOTES = 1              # Number of votes required in CONFIRM_FRAMES for a stable ID.
MAX_BOX_DISTANCE = 50            # Max pixel distance between box centers to consider them the same person.

# --- Global Variables for FAISS ---
# These will hold the FAISS index and the corresponding list of names
FAISS_INDEX = None
# KNOWN_NAMES must be a flattened list, where each entry maps 1:1 to an embedding in the FAISS index.
KNOWN_NAMES = []

# Workbook and sheet for attendance (optional)
workbook = None
sheet = None
try:
    if os.path.exists('DashBoard.xlsx'):
        workbook = openpyxl.load_workbook('DashBoard.xlsx')
        sheet = workbook.active
except Exception:
    workbook = None
    sheet = None

# Runtime set of names already marked/recognized (to avoid duplicate writes)
recognized_names = set()


# --- 1. Model Initialization ---
# Load YOLOv8 model (used for fast, rough face cropping)
try:
    model_yolo = YOLO(YOLO_MODEL_PATH) 
    print("YOLOv8 model loaded successfully.")
except Exception as e:
    print(f"Error loading YOLOv8 model: {e}")
    model_yolo = None

# Load MTCNN (for precise face alignment) and InceptionResnetV1 (FaceNet for embeddings)
try:
    # MTCNN is used to detect and align faces before embedding
    mtcnn = MTCNN(keep_all=True, device='cpu') 
    # ResNet is the FaceNet backbone used to generate the 512-D embedding vector
    resnet = InceptionResnetV1(pretrained='vggface2').eval()
    print("MTCNN and InceptionResnetV1 models loaded successfully.")
except Exception as e:
    print(f"Error loading MTCNN/InceptionResnetV1: {e}")
    mtcnn = None
    resnet = None


# --- 2. Data Loading & Utility Functions ---

def load_known_embeddings() -> Tuple[Optional[object], List[str], bool]:
    """
    Loads known face embeddings and builds a FAISS index for fast nearest-neighbor search.
    FIX: Ensures KNOWN_NAMES is a flattened list corresponding 1:1 with the FAISS index vectors.
    Returns: (faiss_index, list_of_names, success_status)
    """
    global FAISS_INDEX, KNOWN_NAMES
    
    try:
        with open(EMBEDDINGS_FILE, 'rb') as f:
            # raw_embeddings: Dict[str, List[np.ndarray]] = {name: [emb1, emb2, ...]}
            raw_embeddings = pickle.load(f)
            
            if not raw_embeddings:
                print(f"Warning: {EMBEDDINGS_FILE} is empty.")
                return None, [], False

            # --- CRITICAL FIX START ---
            all_names: List[str] = [] # List to hold the name for every single embedding
            all_embeddings: List[np.ndarray] = [] # List to hold every single embedding vector

            for name, embeddings in raw_embeddings.items():
                for emb in embeddings:
                    all_names.append(name)
                    all_embeddings.append(emb)
            
            # 2. Convert to a single NumPy matrix (float32 required by FAISS)
            if not all_embeddings:
                print(f"Warning: No embeddings found after flattening.")
                return None, [], False
                
            embeddings_matrix = np.vstack(all_embeddings).astype('float32')
            D = embeddings_matrix.shape[1] # Dimension (should be 512)
            
            # 3. Initialize FAISS Index (Flat L2 for exact Euclidean distance)
            index = faiss.IndexFlatL2(D)
            index.add(embeddings_matrix)
            
            # --- CRITICAL FIX END ---
            
            print(f"Known embeddings loaded and FAISS index built successfully. ({len(raw_embeddings)} identities, {len(all_names)} total embeddings)")
            
            # Store globally for use in the recognition loop
            FAISS_INDEX = index
            KNOWN_NAMES = all_names # Now KNOWN_NAMES has the same length as the FAISS index
            
            return index, all_names, True
            
    except Exception as e:
        print(f"FATAL: Error loading or building FAISS index from {EMBEDDINGS_FILE}: {e}")
        return None, [], False

def get_face_embedding(face_image: np.ndarray, mtcnn_model: object, resnet_model: object) -> Optional[np.ndarray]:
    """
    Processes a cropped face image to generate a 512-dimensional embedding vector.
    """
    if mtcnn_model is None or resnet_model is None:
        return None
        
    # 1. Convert BGR (OpenCV) to RGB (PyTorch/MTCNN)
    face_rgb = cv2.cvtColor(face_image, cv2.COLOR_BGR2RGB)
    
    # 2. Detect and align the face precisely with MTCNN
    # MTCNN returns a tensor: [N, C, H, W] where N=1 for a single crop
    face_tensor = mtcnn_model(face_rgb)
    
    if face_tensor is None:
        return None # No face detected by MTCNN in the cropped area

    # 3. Generate the embedding vector using FaceNet/InceptionResnetV1
    face_embedding = resnet_model(face_tensor).detach().cpu().numpy().flatten()
    
    # Fix for 1024 vs 512 dimension mismatch during live inference
    expected_dim = 512
    current_dim = face_embedding.shape[0]
    
    if current_dim != expected_dim:
        if current_dim == 1024 and expected_dim == 512:
            # Truncate to 512, assuming the embedding is in the first half
            face_embedding = face_embedding[:expected_dim]
        else:
            print(f"FATAL: Embedding size mismatch. Expected {expected_dim}, got {current_dim}. Cannot proceed.")
            return None
        
    return face_embedding

# --- THREADED CAMERA HANDLER FOR STABILITY (User Provided) ---
class CameraHandler(threading.Thread):
    """
    Handles video capture in a separate thread to prevent the main loop from blocking.
    """
    def __init__(self, stream_url: str):
        super().__init__()
        self.stream_url = stream_url
        self.cap = cv2.VideoCapture(stream_url)
        
        self.cap.set(cv2.CAP_PROP_BUFFERSIZE, 1) 
        
        self.frame = None
        self.success = False
        self.stopped = True
        self.lock = threading.Lock()
        
        if not self.cap.isOpened():
            print(f"Error: Could not open stream {stream_url}. Check URL or camera status.")
        else:
            self.success, self.frame = self.cap.read()
            self.stopped = False
            print("CameraHandler started successfully.")

    def run(self):
        """The main thread loop that continuously reads the newest frame."""
        while not self.stopped:
            success, frame = self.cap.read()
            
            if success:
                with self.lock:
                    self.frame = frame
                    self.success = True
            else:
                self.success = False

    def read(self):
        """Returns the last successfully read frame and its success status."""
        with self.lock:
            return self.success, self.frame

    def stop(self):
        """Signals the thread to stop and releases the video capture."""
        self.stopped = True
        self.join() 
        self.cap.release()
        print("CameraHandler stopped.")


# --- 3. Consistency Helper Functions ---

# Dictionary to store the recognition history and per-track counters for currently tracked faces
# Structure now:
# { face_id: {
#     'box': (x1, y1, x2, y2),
#     'history': ['NameA', 'NameA', 'Unknown', ...],
#     'seen_count': int,             # consecutive frames this face has been seen
#     'analysis_runs': int,          # how many times we've run the costly embedding+search
#     'last_recognition': Optional[str]  # last recognition result from an analysis
#   }
# }
face_history: Dict[int, Dict[str, Any]] = {}
current_face_id: int = 0

def find_nearest_tracked_face(current_box: Tuple[int, int, int, int]) -> Optional[int]:
    """Finds the ID of the closest tracked face from the previous frame."""
    global face_history
    cx_current = (current_box[0] + current_box[2]) / 2
    cy_current = (current_box[1] + current_box[3]) / 2
    
    min_dist = float('inf')
    matched_id = None
    
    for face_id, data in face_history.items():
        hx1, hy1, hx2, hy2 = data['box']
        cx_hist = (hx1 + hx2) / 2
        cy_hist = (hy1 + hy2) / 2
        
        # Calculate Euclidean distance between box centers
        distance = np.sqrt((cx_current - cx_hist)**2 + (cy_current - cy_hist)**2)
        
        if distance < min_dist and distance < MAX_BOX_DISTANCE:
            min_dist = distance
            matched_id = face_id
            
    return matched_id

def write_to_excel(name, status):
    """
    Function to write to the Excel file.
    """
    try:
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            if row[0] == name:
                if name =='none' or name == None or name == 'Unknown' or name == 'Unknown_Unstable':
                    print(f"Skipping write for unrecognized or invalid name: {name}")
                    continue
                # Find the last empty cell in the row
                col_idx = 2  # Start from the second column (Column B)
                while sheet.cell(row=row_idx, column=col_idx).value is not None:
                    col_idx += 1  # Move to the next column

                # Write the status (Late or Absent) with the current date and time
                sheet.cell(row=row_idx, column=col_idx).value = f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} - {status}"
                workbook.save('DashBoard.xlsx')
                print(f"{name} has been recorded as {status} in the Excel file.")
                break
    except Exception as e:
        print(f"Error writing to Excel file: {e}")

def check_absence():
    """
    Function to check and record absence for all names not recognized before 9:00 AM.
    """
    total = 0
    marked = 0
    try:
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True), start=2):
            name = row[0]
            total += 1
            if name not in recognized_names:
                marked += 1
                print(f"{name} is absent.")
                write_to_excel(name, "Absent")
        print(f"Absence recorded for {marked} out of {total} names.")
    except Exception as e:
        print(f"Error during absence recording: {e}")



def determine_stable_name(history_list: List[str]) -> str:
    """Determines the stable name based on the majority vote in the history."""
    if not history_list:
        return "Unknown_Unstable"
        
    counts = {}
    for name in history_list:
        counts[name] = counts.get(name, 0) + 1
        
    # Get the name with the highest count
    stable_name = max(counts, key=counts.get)
    
    if counts[stable_name] >= REQUIRED_VOTES:
        return stable_name
    else:
        # If no single name/identity meets the required votes, it is considered unstable
        return "Unknown_Unstable"


# --- Time helpers ---
def parse_time(s: str) -> dtime:
    """Parse strings like 'HH:MM', 'H' or 'HH' into a datetime.time object.
    Returns midnight (00:00) on parse failure.
    """
    try:
        if s is None:
            return dtime(hour=0, minute=0)
        s_val = str(s).strip()
        if ':' in s_val:
            parts = s_val.split(':')
            h = int(parts[0]) if parts[0] != '' else 0
            m = int(parts[1]) if len(parts) > 1 and parts[1] != '' else 0
        else:
            h = int(s_val)
            m = 0

        h = max(0, min(23, h))
        m = max(0, min(59, m))
        return dtime(hour=h, minute=m)
    except Exception:
        return dtime(hour=0, minute=0)


def is_time_in_range(start: dtime, end: dtime, now_t: dtime) -> bool:
    """Return True if now_t is in the half-open interval [start, end),
    correctly handling intervals that cross midnight."""
    if start <= end:
        return start <= now_t < end
    else:
        # Interval wraps past midnight (e.g., 22:00 -> 02:00)
        return now_t >= start or now_t < end


# --- 4. Real-Time Recognition Loop ---

def real_time_face_recognition():
    global face_history, current_face_id

    # Load time boundaries from config.json (fallbacks provided)
    def load_times(config_path='config.json'):
        defaults = {'start_time': '06:00', 'lateness_time': '07:00', 'absence_time': '10:00'}
        try:
            if os.path.exists(config_path):
                with open(config_path, 'r') as f:
                    cfg = json.load(f)
                    merged = defaults.copy()
                    merged.update({k: v for k, v in cfg.items() if k in merged})
            else:
                merged = defaults
        except Exception:
            merged = defaults

        # Use module-level parse_time helper
        return parse_time(merged.get('start_time')), parse_time(merged.get('lateness_time')), parse_time(merged.get('absence_time'))

    start_time_obj, lateness_time_obj, absence_time_obj = load_times()
    print(f"Attendance window: start={start_time_obj}, lateness={lateness_time_obj}, absence={absence_time_obj}")
    # Track which names we've already printed/recorded to avoid repeats
    marked_names = set()
    # Ensure we reference the module-level recognized_names for absence checking
    global recognized_names

    # Flag to ensure absence is checked only once when the attendance window ends
    absence_checked = False

    # Load the known embeddings and build the FAISS index
    faiss_index, known_names, load_success = load_known_embeddings()
    
    if not load_success or faiss_index is None:
        print("Exiting recognition: FAISS index could not be initialized.")
        return

    # Initialize Camera Handler (IP camera feed)
    camera = CameraHandler(STREAM_URL)
    if camera.stopped: # Check if initialization failed
        return
    now = datetime.now().time()
    # Only start the camera loop if the current time is within the attendance window
    if is_time_in_range(start_time_obj, absence_time_obj, now):
        camera.start()
        # Schedule an automatic absence check at the next occurrence of absence_time
        try:
            now_dt = datetime.now()
            next_absence_dt = datetime.combine(now_dt.date(), absence_time_obj)
            if next_absence_dt <= now_dt:
                next_absence_dt += timedelta(days=1)
            delay_seconds = (next_absence_dt - now_dt).total_seconds()

            def absence_action():
                try:
                    if sheet is not None:
                        check_absence()
                        print("Scheduled absence check complete.")
                except Exception as e:
                    print(f"Error in scheduled absence action: {e}")
                # Terminate the process after recording absences to avoid repeated runs
                try:
                    camera.stop()
                except Exception:
                    pass
                try:
                    cv2.destroyAllWindows()
                except Exception:
                    pass
                # Exit now
                os._exit(0)

            t = threading.Timer(delay_seconds, absence_action)
            t.daemon = True
            t.start()
        except Exception as e:
            print(f"Warning: failed to schedule absence check: {e}")
    else:
        print("Outside of attendance marking hours. Exiting recognition.")
        return
    
    try:
        print("\n--- Starting Real-Time Recognition (Press 'q' to exit) ---")
        
        while True:
            success, frame = camera.read()
            
            if not success or frame is None:
                time.sleep(0.01)
                continue

            # If the attendance window has ended, run absence checks once and exit
            now = datetime.now().time()
            if (not absence_checked) and (not is_time_in_range(start_time_obj, absence_time_obj, now)):
                try:
                    if sheet is not None:
                        check_absence()
                except Exception as e:
                    print(f"Error running absence check: {e}")
                absence_checked = True
                print("Attendance window ended — recorded absences and exiting.")
                break

            # 1. Use YOLO to detect faces in the current frame
            yolo_results = model_yolo(frame, verbose=False, conf=YOLO_CONFIDENCE_THRESHOLD) 
            
            # Temporary dict to hold associations for the current frame
            current_frame_detections: Dict[int, Dict[str, Any]] = {}
            
            for box in yolo_results[0].boxes:
                # Get bounding box coordinates
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                current_box = (x1, y1, x2, y2)
                
                # --- Tracking and History Lookup ---
                matched_id = find_nearest_tracked_face(current_box)
                
                if matched_id is None:
                    # New face detected: assign a new ID
                    matched_id = current_face_id
                    current_face_id += 1
                    # Initialize tracking info for the new face
                    history_list = []
                    seen_count = 0
                    analysis_runs = 0
                    last_recognition = None
                else:
                    # Existing face: retrieve its history and counters
                    tracked = face_history.get(matched_id, {})
                    history_list = tracked.get('history', [])
                    seen_count = tracked.get('seen_count', 0)
                    analysis_runs = tracked.get('analysis_runs', 0)
                    last_recognition = tracked.get('last_recognition', None)
                
                # Update the temporary tracker for the next iteration
                # We'll store additional per-track counters in the temp structure
                current_frame_detections[matched_id] = {
                    'box': current_box,
                    'history': history_list,
                    'seen_count': seen_count,
                    'analysis_runs': analysis_runs,
                    'last_recognition': last_recognition
                }
                # Per-track seen counter increases because we have a match in this frame
                current_frame_detections[matched_id]['seen_count'] = seen_count + 1

                # Decide whether to run the costly embedding + FAISS search on THIS frame
                # Behavior requested:
                # - Analyze every frame while seen_count < 3 (fast warm-up)
                # - Once seen_count >= 3, allow up to two analysis runs: first with base threshold,
                #   then one more with an INCREASED threshold. After those, skip analysis and
                #   reuse the last recognition to fill the history (votes still required).
                recognition_result = "Unknown"  # default for this single frame
                min_distance = float('inf')
                best_match = None

                # Calculate whether we should analyze this frame
                sc = current_frame_detections[matched_id]['seen_count']
                ar = current_frame_detections[matched_id]['analysis_runs']
                lr = current_frame_detections[matched_id]['last_recognition']

                # Threshold increment used for the second analysis
                THRESHOLD_INCREMENT = 0.2

                analyze_now = False
                used_threshold = RECOGNITION_THRESHOLD

                if sc < 3:
                    # Warm-up: analyze every frame
                    analyze_now = True
                    used_threshold = RECOGNITION_THRESHOLD
                else:
                    # sc >= 3: allow up to two analysis runs
                    if ar == 0:
                        analyze_now = True
                        used_threshold = RECOGNITION_THRESHOLD  # first post-3 analysis uses base threshold
                    elif ar == 1:
                        analyze_now = True
                        used_threshold = RECOGNITION_THRESHOLD + THRESHOLD_INCREMENT  # second analysis uses increased threshold
                    else:
                        analyze_now = False

                face_crop = frame[y1:y2, x1:x2]

                if analyze_now:
                    embedding = get_face_embedding(face_crop, mtcnn, resnet)
                    if embedding is not None:
                        # Reshape for FAISS
                        query_embedding = embedding.reshape(1, -1).astype('float32')
                        K = 1
                        D, I = faiss_index.search(query_embedding, K)
                        min_distance = D[0][0]
                        best_match_index = I[0][0]
                        if best_match_index != -1:
                            best_match = known_names[best_match_index]
                        else:
                            best_match = None

                        # Compare against the (possibly increased) threshold
                        if best_match is not None and min_distance < (used_threshold ** 2):
                            recognition_result = best_match
                        else:
                            recognition_result = "Unknown"
                    else:
                        recognition_result = "Unknown"

                    # Update analysis_runs and last_recognition for this track
                    current_frame_detections[matched_id]['analysis_runs'] = ar + 1
                    current_frame_detections[matched_id]['last_recognition'] = recognition_result
                else:
                    # Not analyzing this frame: reuse the last recognition result (if any)
                    if lr is not None:
                        recognition_result = lr
                    else:
                        recognition_result = "Unknown"
                
                # Update the history with the single-frame recognition result
                tracked_entry = current_frame_detections[matched_id]
                hist = tracked_entry.get('history', [])
                hist.append(recognition_result)
                # Keep history bounded
                if len(hist) > CONFIRM_FRAMES:
                    hist.pop(0)

                # Write back history in case it's a new list
                tracked_entry['history'] = hist

                # 5. Determine the STABLE name based on the history (The "Double Check")
                stable_name = determine_stable_name(hist)
                display_name = ""

                if stable_name != "Unknown_Unstable":
                    # Stable match found (green box)
                    display_name = f"{stable_name} (Stable)"
                    color = (0, 255, 0)

                    # Determine current status based on loaded schedule
                    now = datetime.now().time()
                    status = None
                    if is_time_in_range(start_time_obj, lateness_time_obj, now):
                        status = 'Present'
                    elif is_time_in_range(lateness_time_obj, absence_time_obj, now):
                        status = 'Late'
                    else:
                        status = 'absent'


                    # Print and record the first time we see this stable name
                    if stable_name not in marked_names:
                        print(f"{stable_name} - {status}")
                        try:
                            if sheet is not None:
                                write_to_excel(stable_name, status)
                        except Exception:
                            print(f"Warning: failed to write {stable_name} - {status} to Excel")
                        marked_names.add(stable_name)
                        try:
                            recognized_names.add(stable_name)
                        except Exception:
                            pass
                # 6. Draw the bounding box and name on the frame
                cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)
                cv2.putText(frame, display_name, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.7, color, 2)
            
            # Update the global history for the next frame
            face_history = current_frame_detections
            
            # Display the result
            cv2.imshow("Real-Time Face Recognition", frame)
            
            # Exit loop if 'q' is pressed
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break

    finally:
        # Cleanup
        camera.stop()
        cv2.destroyAllWindows()


# --- 5. Main Execution Block ---

if __name__ == "__main__":
    # Ensure FAISS is available
    try:
        import faiss
    except ImportError:
        print("FATAL ERROR: FAISS is not installed. Please install it using: pip install faiss-cpu")
        exit()
    if model_yolo is None or mtcnn is None or resnet is None:
        print("\nExiting: Critical models failed to load. Please check model paths and dependencies.")
    else:
        # Run the real-time recognition loop
        real_time_face_recognition()

    print("\nApplication terminated.")
